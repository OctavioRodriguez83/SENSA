import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from abv_web.models import Categoria

class Command(BaseCommand):
    help = 'Importa categorías desde un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--path', '-p', type=str, help='Ruta al archivo .xlsx (opcional).')

    def handle(self, *args, **options):
        ruta = options['path']
        if not ruta or not os.path.isfile(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        self.stdout.write(f"Leyendo Excel desde: {ruta}")
        wb = load_workbook(filename=ruta, data_only=True)
        try:
            sheet = wb['Lista Categorias']
        except KeyError:
            self.stdout.write(self.style.WARNING('La hoja "Lista Categorias" no existe. Usando la primera hoja.'))
            sheet = wb[wb.sheetnames[0]]  # Selecciona la primera hoja

        # Validar encabezados
        headers = [cell.value for cell in sheet[1]]
        if headers[:3] != ['Nombre de la categoría', 'Descripcion', 'Imagen']:
            raise CommandError("El archivo Excel no tiene los encabezados esperados: 'Nombre de la categoría', 'Descripcion', 'Imagen'.")

        categorias_creadas = 0
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # Detener si la fila está completamente vacía
            if all(cell.value is None for cell in row):
                break

            categoria_name, categoria_descripcion, categoria_url_img = [cell.value for cell in row[:3]]

            if not categoria_name or not categoria_descripcion:
                self.stdout.write(self.style.WARNING(
                    f"Fila {idx}: 'Nombre de la categoría' o 'Descripcion' está vacío. Saltando fila."
                ))
                continue

            # Si hay imagen definida, se cambia el nombre agregando el prefijo Imagenes_
            imagen_filename = None
            if categoria_url_img:
                categoria_url_img_str = str(categoria_url_img).strip()
                _, ext = os.path.splitext(categoria_url_img_str)
                imagen_filename = f"Imagenes_Categoria/{categoria_url_img_str}"

            Categoria.objects.update_or_create(
                categoria_name=categoria_name.strip(),
                defaults={
                    'categoria_url_img': imagen_filename,
                    'categoria_descripcion': categoria_descripcion.strip(),
                    'statusCategoria': True,  # Inyecta True por defecto
                }
            )
            categorias_creadas += 1

        self.stdout.write(self.style.SUCCESS(f"✅ Se importaron {categorias_creadas} categorías correctamente."))