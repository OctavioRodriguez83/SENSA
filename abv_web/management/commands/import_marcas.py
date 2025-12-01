import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from abv_web.models import Marca

class Command(BaseCommand):
    help = 'Importa marcas desde un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--path', '-p', type=str,
                            help='Ruta al archivo .xlsx (por defecto `files/importaciones/marcas.xlsx`).')

    def handle(self, *args, **options):
        ruta = options['path']
        if not ruta or not os.path.isfile(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        self.stdout.write(f"Leyendo Excel desde: {ruta}")
        wb = load_workbook(filename=ruta, data_only=True)
        try:
            sheet = wb['Lista Marcas']
        except KeyError:
            self.stdout.write(self.style.WARNING('La hoja "Lista Marcas" no existe. Usando la primera hoja.'))
            sheet = wb[wb.sheetnames[0]]  # Selecciona la primera hoja

        # Validar encabezados
        headers = [cell.value for cell in sheet[1]]
        if headers[:3] != ['Nombre de la marca', 'Descripcion', 'Imagen']:
            raise CommandError("El archivo Excel no tiene los encabezados esperados: 'Nombre de la marca', 'Descripcion', 'Imagen'.")

        marcas_creadas = 0
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            marca_name, marca_descripcion, marca_url_img = [cell.value for cell in row[:3]]

            if not marca_name or not marca_descripcion:
                self.stdout.write(self.style.WARNING(
                    f"Fila {idx}: 'Nombre de la marca' o 'Descripcion' está vacío. Saltando fila."
                ))
                continue

            # Si hay imagen definida, se cambia el nombre agregando el prefijo Imagenes_
            imagen_filename = None
            if marca_url_img:
                marca_url_img_str = str(marca_url_img).strip()
                _, ext = os.path.splitext(marca_url_img_str)
                imagen_filename = f"Imagenes_Marca/{marca_url_img_str}"

            Marca.objects.update_or_create(
                marca_name=marca_name.strip(),
                defaults={
                    'marca_url_img': imagen_filename,
                    'marca_descripcion': marca_descripcion.strip(),
                    'status': True,  # Inyecta True por defecto
                }
            )
            marcas_creadas += 1

        self.stdout.write(self.style.SUCCESS(f"✅ Se importaron {marcas_creadas} marcas correctamente."))