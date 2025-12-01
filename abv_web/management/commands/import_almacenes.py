import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from abv_web.models import Empresa, Almacen

class Command(BaseCommand):
    help = 'Importa almacenes desde un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--path', '-p', type=str, help='Ruta al archivo .xlsx (opcional).')

    def handle(self, *args, **options):
        ruta = options['path']
        if not ruta or not os.path.isfile(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        self.stdout.write(f"Leyendo Excel desde: {ruta}")
        wb = load_workbook(filename=ruta, data_only=True)
        try:
            sheet = wb['Lista Almacenes']
        except KeyError:
            raise CommandError('La hoja "Lista Almacenes" no existe en el archivo.')

        # Validar encabezados
        headers = [cell.value for cell in sheet[1]]
        esperados = [
            'Nombre del almacen', 'Calle', 'Numero', 'Colonia',
            'Codigo Postal', 'Ciudad', 'Estado', 'Empresa'
        ]
        if headers[:8] != esperados:
            raise CommandError(
                "El archivo Excel no tiene los encabezados esperados: "
                + ", ".join(repr(h) for h in esperados) + "."
            )

        # Mapa de empresas (clave: nombre en minúsculas y sin espacios extras)
        empresa_map = {
            str(e.empresa_nombre).strip().lower(): e.id
            for e in Empresa.objects.all()
        }

        not_inserted = []
        almacenes_creados = 0

        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # Si la fila está vacía, cortamos
            if all(cell.value is None for cell in row):
                break

            # Convertir todo a str y strip para evitar errores con int, None, etc.
            almacen_nombre = str(row[0].value or '').strip()
            calle          = str(row[1].value or '').strip()
            numero         = str(row[2].value or '').strip()
            colonia        = str(row[3].value or '').strip()
            codigo_postal  = str(row[4].value or '').strip()
            ciudad         = str(row[5].value or '').strip()
            estado         = str(row[6].value or '').strip()
            empresa_nombre = str(row[7].value or '').strip().lower()

            # Validar campos obligatorios
            if not all([almacen_nombre, calle, numero, colonia,
                        codigo_postal, ciudad, estado, empresa_nombre]):
                not_inserted.append({
                    'fila': idx,
                    'razon': 'Campos obligatorios vacíos'
                })
                continue

            # Buscar empresa
            empresa_id = empresa_map.get(empresa_nombre)
            if not empresa_id:
                not_inserted.append({
                    'fila': idx,
                    'razon': f"Empresa '{row[7].value}' no encontrada"
                })
                continue

            # Crear o actualizar
            Almacen.objects.update_or_create(
                almacen_nombre=almacen_nombre,
                defaults={
                    'almacen_direccion_calle': calle,
                    'almacen_direccion_numero': numero,
                    'almacen_direccion_colonia': colonia,
                    'almacen_direccion_cp': codigo_postal,
                    'almacen_direccion_ciudad': ciudad,
                    'almacen_direccion_estado': estado,
                    'empresa_id': empresa_id,
                }
            )
            almacenes_creados += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"✅ Se importaron {almacenes_creados} almacenes correctamente."
            )
        )

        # Guardar log de no insertados
        if not_inserted:
            log_path = os.path.join('files', 'importaciones', 'almacenes_no_registrados.md')
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            with open(log_path, 'w', encoding='utf-8') as f:
                for item in not_inserted:
                    f.write(f"- Fila: {item['fila']} | Razón: {item['razon']}\n")
            self.stdout.write(
                self.style.WARNING(
                    f"Almacenes no registrados guardados en `{log_path}`"
                )
            )
