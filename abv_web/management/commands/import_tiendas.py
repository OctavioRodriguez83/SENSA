import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from abv_web.models import Tienda

class Command(BaseCommand):
    help = 'Importa tiendas desde un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--path', '-p', type=str,
                            help='Ruta al archivo .xlsx (por defecto `files/importaciones/tiendas.xlsx`).')

    def handle(self, *args, **options):
        ruta = options['path']
        if not ruta or not os.path.isfile(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        self.stdout.write(f"Leyendo Excel desde: {ruta}")
        wb = load_workbook(filename=ruta, data_only=True)
        try:
            sheet = wb['Lista Tiendas']
        except KeyError:
            self.stdout.write(self.style.WARNING('La hoja "Lista Tiendas" no existe. Usando la primera hoja.'))
            sheet = wb[wb.sheetnames[0]]  # Selecciona la primera hoja

        # Validar encabezados
        headers = [cell.value for cell in sheet[1]]
        if headers[:4] != ['Enlace de la tienda (Link de sitio web)', 'Nombre de la tienda', 'Descripcion de la tienda', 'Imagen']:
            raise CommandError("El archivo Excel no tiene los encabezados esperados: 'Enlace de la tienda (Link de sitio web)', 'Nombre de la tienda', 'Descripcion de la tienda', 'Imagen'.")

        tiendas_creadas = 0
        not_inserted = []

        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            enlace, nombre, descripcion, imagen = [cell.value for cell in row[:4]]

            # Validar campos obligatorios
            if not enlace or not nombre or not descripcion:
                not_inserted.append({'fila': idx, 'razon': 'Campos obligatorios vacíos'})
                continue

            # Se modifica el nombre de la imagen con el prefijo Imagenes_ y el nombre de la tienda
            imagen_filename = None
            if imagen:
                imagen_str = str(imagen).strip()
                _, ext = os.path.splitext(imagen_str)
                imagen_filename = f"Imagenes_Tienda/{imagen_str}"

            # Crear o actualizar la tienda
            Tienda.objects.update_or_create(
                tienda_nombre=nombre.strip(),
                defaults={
                    'tienda_enlace': enlace.strip(),
                    'tienda_descripcion': descripcion.strip(),
                    'tienda_imagen': imagen_filename,
                }
            )
            tiendas_creadas += 1

        self.stdout.write(self.style.SUCCESS(f"✅ Se importaron {tiendas_creadas} tiendas correctamente."))

        # Guardar errores en un archivo de log
        if not_inserted:
            log_path = os.path.join('files', 'importaciones', 'tiendas_no_registradas.md')
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            with open(log_path, 'w', encoding='utf-8') as f:
                for item in not_inserted:
                    f.write(f"- Fila: {item['fila']} | Razón: {item['razon']}\n")
            self.stdout.write(self.style.WARNING(f"Tiendas no registradas guardadas en `{log_path}`"))