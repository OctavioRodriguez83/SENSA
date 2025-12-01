
import os
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from abv_web.models import Empresa

class Command(BaseCommand):
    help = 'Importa empresas desde un archivo Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--path', '-p', type=str, help='Ruta al archivo .xlsx (opcional).')

    def handle(self, *args, **options):
        ruta = options['path']
        if not ruta or not os.path.isfile(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        self.stdout.write(f"Leyendo Excel desde: {ruta}")
        wb = load_workbook(filename=ruta, data_only=True)
        try:
            sheet = wb['Lista Empresas']
        except KeyError:
            self.stdout.write(self.style.WARNING('La hoja "Lista Empresas" no existe. Usando la primera hoja.'))
            sheet = wb[wb.sheetnames[0]]  # Selecciona la primera hoja

        # Validar encabezados
        headers = [cell.value for cell in sheet[1]]
        if headers[:3] != ['Nombre de la empresa', 'Descripcion', 'Logo']:
            raise CommandError("El archivo Excel no tiene los encabezados esperados: 'Nombre de la empresa', 'Descripcion', 'Logo'.")

        empresas_creadas = 0
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            # Detener si la fila está completamente vacía
            if all(cell.value is None for cell in row):
                break

            empresa_nombre, empresa_descripcion, empresa_logo = [cell.value for cell in row[:3]]

            if not empresa_nombre or not empresa_descripcion:
                self.stdout.write(self.style.WARNING(
                    f"Fila {idx}: 'Nombre de la empresa' o 'Descripcion' está vacío. Saltando fila."
                ))
                continue

            # Si hay logo definido, se cambia el nombre agregando el prefijo Imagenes_
            logo_filename = None
            if empresa_logo:
                empresa_logo_str = str(empresa_logo).strip()
                nombre_base, ext = os.path.splitext(empresa_logo_str)
                logo_filename = f"Imagenes_Empresa/{empresa_logo_str}"

            Empresa.objects.update_or_create(
                empresa_nombre=empresa_nombre.strip(),
                defaults={
                    'empresa_logo': logo_filename,
                    'empresa_descripcion': empresa_descripcion.strip(),
                }
            )
            empresas_creadas += 1

        self.stdout.write(self.style.SUCCESS(f"✅ Se importaron {empresas_creadas} empresas correctamente."))