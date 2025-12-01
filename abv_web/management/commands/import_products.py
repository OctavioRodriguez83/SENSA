import os
import warnings
from django.core.management.base import BaseCommand, CommandError
from django.db import connection, transaction
from openpyxl import load_workbook
from abv_web.models import Marca, Categoria

TABLE_NAME = 'abv_web_producto'
COLUMNS = [
    'producto_extendido',
    'producto_sku',
    'producto_skunetsuite',
    'producto_ean',
    'producto_nombre',
    'producto_descripcion',
    'producto_modelo',
    'producto_precio_base',
    'producto_precio_amazon',
    'producto_precio_mercadolibre',
    'producto_precio_ebay',
    'producto_url_img',
    'marca_id',
    'categoria_id',
]

class Command(BaseCommand):
    help = 'Importa productos desde un archivo Excel cargado.'

    def add_arguments(self, parser):
        parser.add_argument('--path', '-p', type=str, help='Ruta al archivo .xlsx (opcional).')
        parser.add_argument('--file', type=str, help='Archivo temporal cargado (opcional).')

    def handle(self, *args, **options):
        ruta = options['path']
        archivo_temporal = options['file']

        if not ruta and not archivo_temporal:
            raise CommandError("Debes proporcionar una ruta o un archivo temporal.")

        if archivo_temporal:
            ruta = archivo_temporal

        if not os.path.isfile(ruta):
            raise CommandError(f"No se encontró el archivo: {ruta}")

        self.stdout.write(f"Leyendo Excel desde: {ruta}")
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
        wb = load_workbook(filename=ruta, data_only=True)
        try:
            sheet = wb['Lista Productos']
        except KeyError:
            raise CommandError('La hoja "Lista Productos" no existe en el archivo.')

        marca_map = {m.marca_name.strip().lower(): m.id for m in Marca.objects.all()}
        categoria_map = {c.categoria_name.strip().lower(): c.id for c in Categoria.objects.all()}

        not_inserted = []
        rows = []
        for idx, excel_row in enumerate(sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in excel_row[:14]]

            producto_sku, producto_skunetsuite, producto_ean, producto_nombre = values[1:5]
            if any(val is None or str(val).strip() == '' for val in [producto_sku, producto_skunetsuite, producto_ean, producto_nombre]):
                not_inserted.append({'sku': producto_sku, 'razon': 'Campos obligatorios vacíos'})
                continue

            nombre_marca = values[12]
            marca_id = marca_map.get(nombre_marca.strip().lower()) if nombre_marca else None

            nombre_categoria = values[13]
            categoria_id = categoria_map.get(nombre_categoria.strip().lower()) if nombre_categoria else None

            # Se modifica el nombre de la imagen con el prefijo Imagenes_ y el nombre del producto
            if values[11]:
                prod_img_str = str(values[11]).strip()
                _, ext = os.path.splitext(prod_img_str)
                producto_url_img = f"Imagenes_Producto/{prod_img_str}"
            else:
                producto_url_img = None

            rows.append([
                values[0], producto_sku, producto_skunetsuite, producto_ean, producto_nombre,
                values[5], values[6], values[7] or 0, values[8] or 0, values[9] or 0,
                values[10] or 0, producto_url_img, marca_id, categoria_id,
            ])

        if not rows:
            self.stdout.write(self.style.ERROR('No se encontraron filas para importar.'))
        else:
            placeholders = ','.join(['%s'] * len(COLUMNS))
            column_list = ','.join(COLUMNS)
            sql = f"INSERT INTO {TABLE_NAME} ({column_list}) VALUES ({placeholders});"
            with transaction.atomic():
                with connection.cursor() as cursor:
                    cursor.executemany(sql, rows)
            self.stdout.write(self.style.SUCCESS(f"✅ Insertadas {len(rows)} filas en {TABLE_NAME}."))

        if not_inserted:
            log_path = os.path.join('files', 'importaciones', 'productos_no_registrados.md')
            with open(log_path, 'w', encoding='utf-8') as f:
                for item in not_inserted:
                    f.write(f"- SKU: {item['sku']} | Razón: {item['razon']}\n")
            self.stdout.write(self.style.WARNING(f"Productos no registrados guardados en `{log_path}`"))